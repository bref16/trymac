# -*- coding: utf-8 -*-
import sys, re, os, subprocess
from dataclasses import dataclass
from typing import Dict, Optional, Tuple, List
from functools import partial
from datetime import datetime

from PySide6 import QtWidgets, QtCore
from PySide6.QtWidgets import QDialog, QFormLayout, QDialogButtonBox
from sqlalchemy import create_engine, text, inspect
from sqlalchemy.engine import Engine

try:
    import openpyxl
except Exception:
    openpyxl = None

LABEL_TO_TABLE = {
    "Аппарат ИВЛ": "Block_Main",
    "Лицензия": "License",
    "Контуры": "Circuits",
    "Клапаны, датчики": "Valves",
    "Маски": "Masks",
    "Мобильная стойка": "Mobile_Cart",
    "Автокрепления": "Holders",
    "Увлажнитель": "Humidifier",
    "Датчик CO2": "CO2",
    "Датчик SpO2": "O2",
}
TIN_ALL_TABLE = 'EVE TIN ALL'
TEMPLATES_TABLE = 'Templates'
MODES_TABLE = 'Modes'
MODES_COL = 'Mode'

ORDERED_LABELS = [
    "Аппарат ИВЛ","Лицензия","Контуры","Клапаны, датчики",
    "Маски","Мобильная стойка","Автокрепления","Увлажнитель",
    "Датчик CO2","Датчик SpO2"
]

def qident(name: str) -> str:
    return '"' + name.replace('"','""') + '"'

def has_columns(engine: Engine, table: str, cols: List[str]) -> Dict[str, bool]:
    insp = inspect(engine)
    existing = {c['name'] for c in insp.get_columns(table)}
    return {c: (c in existing) for c in cols}

def norm_ref(x) -> str:
    if x is None: return ""
    s = str(x).strip().replace(",", ".")
    m = re.match(r"^(\d+)\.0$", s)
    if m: return m.group(1)
    try:
        if re.match(r"^\d+(\.\d+)?$", s):
            f = float(s)
            if f.is_integer(): return str(int(f))
    except Exception:
        pass
    return s

def digits_only(s: str) -> str:
    return re.sub(r"\D", "", s or "")

def parse_money(s: Optional[str]) -> Optional[float]:
    if s is None: return None
    t = str(s).strip().replace(",", ".")
    try:
        return float(t)
    except Exception:
        return None

def fmt_money(x: Optional[float]) -> str:
    if x is None: return ""
    return f"{x:.2f}"

def app_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

@dataclass
class RowControl:
    label: str
    combo: QtWidgets.QComboBox
    spin: QtWidgets.QSpinBox

class Panel(QtWidgets.QWidget):
    COL_NO = 0
    COL_PN = 1
    COL_DESC = 2
    COL_QTY = 3
    COL_CTRL = 4
    COL_PACK = 5
    COL_PRICE_LIST = 6
    COL_PRICE_TRIMM = 7

    USERROLE_BASE_TRIMM = QtCore.Qt.UserRole + 1
    USERROLE_LIST_PRICE = QtCore.Qt.UserRole + 2

    def __init__(self):
        super().__init__()
        self.setWindowTitle("MAG Config — DB")
        self.setMinimumWidth(1400)

        self.engine: Optional[Engine] = None
        self.inspector = None

        self.available_modes: List[str] = ["EVE", "S", "F"]
        self.current_mode = self.available_modes[0]
        self.side_filter: Optional[str] = None

        self.labels: List[str] = ORDERED_LABELS[:]
        self.row_controls: List[RowControl] = []

        self.table_cache: Dict[str, List[Dict[str,str]]] = {}
        self.options_cache: Dict[Tuple[str, str, Optional[str]], Dict[str, str]] = {}
        self.tin_index: Dict[str, Tuple[str, Optional[str], Optional[str], Optional[str]]] = {}

        self.tin_ref_col = None
        self.tin_desc_col = None
        self.tin_pack_col = None
        self.tin_price_list_col = None
        self.tin_price_trimm_col = None

        self.templates: Dict[str, List[Tuple[str, int]]] = {}

        self._building_table = False
        self._build_ui()

    # ================= UI =================
    def _build_ui(self):
        # hidden connection fields
        self.hostEdit = QtWidgets.QLineEdit("127.0.0.1")
        self.portEdit = QtWidgets.QLineEdit("5432")
        self.dbEdit   = QtWidgets.QLineEdit("mag_config")
        self.userEdit = QtWidgets.QLineEdit("postgres")
        self.pwEdit   = QtWidgets.QLineEdit(""); self.pwEdit.setEchoMode(QtWidgets.QLineEdit.Password)

        # -------- Header (single line): modes on left, admin+connect on right
        self.modeBar = QtWidgets.QWidget()
        self.modeBarLayout = QtWidgets.QHBoxLayout(self.modeBar)
        self.modeBarLayout.setContentsMargins(0, 0, 0, 0)
        self._rebuild_mode_strip()

        self.btnAdmin = QtWidgets.QPushButton("Редактировать значения/таблицы")
        self.btnAdmin.clicked.connect(self._launch_admin_tool)
        self.btnConn = QtWidgets.QPushButton("Настроить подключение к БД")
        self.btnConn.clicked.connect(self._show_conn_dialog)

        header = QtWidgets.QHBoxLayout()
        header.addWidget(self.modeBar)
        header.addStretch(1)
        header.addWidget(self.btnAdmin)
        header.addSpacing(8)
        header.addWidget(self.btnConn)

        # -------- Left, Right panes
        self.leftBox = QtWidgets.QGroupBox("Блоки (слева)")
        self.leftLayout = QtWidgets.QVBoxLayout(self.leftBox)

        self.actionsBox = self._make_actions_box()
        self.summaryBox = QtWidgets.QGroupBox("Сводка")
        rlay = QtWidgets.QVBoxLayout(self.summaryBox)

        self.table = QtWidgets.QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels([
            "№","Кат. №","Описаное","К-во","Control","шт/уп","Стоимость","Цена ТРИММ"
        ])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.itemChanged.connect(self._on_table_item_changed)
        rlay.addWidget(self.table)

        row_ops = QtWidgets.QHBoxLayout()
        self.btn_clear_cat = QtWidgets.QPushButton("Очистить Кат.№")
        self.btn_move_down  = QtWidgets.QPushButton("Сместить вниз строку с Кат№")
        self.btn_delete     = QtWidgets.QPushButton("Удалить строку")
        for b in (self.btn_clear_cat, self.btn_move_down, self.btn_delete):
            b.setAutoDefault(False); b.setDefault(False)
        self.btn_clear_cat.clicked.connect(self._clear_cat_numbers)
        self.btn_move_down.clicked.connect(self._move_row_with_cat_down)
        self.btn_delete.clicked.connect(self.delete_selected)
        row_ops.addWidget(self.btn_clear_cat)
        row_ops.addWidget(self.btn_move_down)
        row_ops.addWidget(self.btn_delete)
        row_ops.addStretch(1)
        rlay.addLayout(row_ops)

        controls = QtWidgets.QHBoxLayout()
        self.discountSpin = QtWidgets.QDoubleSpinBox()
        self.discountSpin.setRange(0.0, 100.0); self.discountSpin.setDecimals(2); self.discountSpin.setSuffix(" %")
        self.discountSpin.setValue(0.0)

        self.logisticsSpin = QtWidgets.QDoubleSpinBox()
        self.logisticsSpin.setRange(0.0, 1000.0); self.logisticsSpin.setDecimals(4)
        self.logisticsSpin.setValue(1.0000)

        self.kursSpin = QtWidgets.QDoubleSpinBox()
        self.kursSpin.setRange(0.0, 100000.0); self.kursSpin.setDecimals(4)
        self.kursSpin.setValue(1.0000)

        for w, lbl in ((self.discountSpin, "Скидка %"),
                       (self.logisticsSpin, "Логистика"),
                       (self.kursSpin, "Курс")):
            controls.addWidget(QtWidgets.QLabel(lbl))
            controls.addWidget(w)
            controls.addSpacing(12)
        controls.addStretch(1)
        rlay.addLayout(controls)

        totals = QtWidgets.QHBoxLayout()
        self.lbl_total = QtWidgets.QLabel("Итого: 0.00")
        self.lbl_margin = QtWidgets.QLabel("Маржа: —")
        fontB = self.lbl_total.font(); fontB.setBold(True)
        self.lbl_total.setFont(fontB); self.lbl_margin.setFont(fontB)
        totals.addWidget(self.lbl_total); totals.addSpacing(24)
        totals.addWidget(self.lbl_margin); totals.addStretch(1)
        rlay.addLayout(totals)

        rightPane = QtWidgets.QWidget()
        right_v = QtWidgets.QVBoxLayout(rightPane)
        right_v.addWidget(self.actionsBox)
        right_v.addWidget(self.summaryBox, 1)

        split = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        split.addWidget(self.leftBox)
        split.addWidget(rightPane)
        split.setStretchFactor(0, 2)
        split.setStretchFactor(1, 1)

        self.status = QtWidgets.QLabel("Нажмите «Настроить подключение к БД» для подключения к базе.")

        main = QtWidgets.QVBoxLayout(self)
        main.addLayout(header)
        main.addWidget(split, 1)
        main.addWidget(self.status)

    def _rebuild_mode_strip(self):
        while self.modeBarLayout.count():
            item = self.modeBarLayout.takeAt(0)
            w = item.widget()
            if w: w.deleteLater()

        modes = [m for m in self.available_modes if str(m).strip()]
        if not modes:
            self.modeBarLayout.addWidget(QtWidgets.QLabel("Режимы: (подключитесь к БД)"))
            self.modeBarLayout.addStretch(1)
            return

        for m in modes:
            b = QtWidgets.QPushButton(str(m))
            b.setMinimumWidth(90); b.setMinimumHeight(34)
            b.setAutoDefault(False); b.setDefault(False)
            b.clicked.connect(lambda _, mm=str(m): self.set_mode(mm))
            self.modeBarLayout.addWidget(b)

    def _make_actions_box(self) -> QtWidgets.QGroupBox:
        box = QtWidgets.QGroupBox("Образцы КП / Действия")
        grid = QtWidgets.QGridLayout(box)
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(8)
        grid.setContentsMargins(8, 8, 8, 8)
        box.setSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Fixed)

        def add_btn(text, r, c, slot):
            b = QtWidgets.QPushButton(text)
            b.setMinimumHeight(28)
            b.setAutoDefault(False); b.setDefault(False)
            b.clicked.connect(slot)
            grid.addWidget(b, r, c)
            return b

        # template buttons
        add_btn("EVE TR", 0, 0, lambda: self._on_template_button("EVE TR", mode="EVE"))
        add_btn("S",      0, 1, lambda: self._on_template_button("S", mode="S"))
        add_btn("F прав", 0, 2, lambda: self._on_template_button("F прав", mode="F", side="прав"))
        # quick connect inside the grid (extra column)
        self.btnQuickConnect = add_btn("Подключиться", 0, 3, self._quick_connect)

        add_btn("EVE NEO", 1, 0, lambda: self._on_template_button("EVE NEO", mode="EVE"))
        add_btn("F лев",   1, 1, lambda: self._on_template_button("F лев", mode="F", side="лев"))
        add_btn("EVE IN",  2, 0, lambda: self._on_template_button("EVE IN", mode="EVE"))
        add_btn("EVE ALL", 3, 0, lambda: self._on_template_button("EVE ALL", mode="EVE"))
        add_btn("Выгрузить КП", 4, 0, self._export_kp)
        add_btn("КП ИВЛ", 5, 0, self._kp_ivl)
        add_btn("КП расх", 5, 1, self._kp_raskh)
        add_btn("PO Stephan", 6, 0, self._po_stephan)

        spacer = QtWidgets.QWidget(); spacer.setFixedHeight(6)
        grid.addWidget(spacer, 7, 0, 1, 4)
        return box

    # ============== DB & data ==============
    def _make_url(self) -> str:
        return f'postgresql+psycopg://{self.userEdit.text()}:{self.pwEdit.text()}@{self.hostEdit.text()}:{self.portEdit.text()}/{self.dbEdit.text()}'

    def connect_db(self) -> bool:
        try:
            url = self._make_url()
            self.engine = create_engine(url, pool_pre_ping=True)
            with self.engine.connect() as conn:
                conn.execute(text("SELECT 1"))
            self.inspector = inspect(self.engine)
            self.status.setText("Соединение установлено, загрузка данных…")

            self._load_modes()
            self._detect_tin_columns()
            self._preload_all_db()
            self._preload_templates()

            self._rebuild_mode_strip()
            if self.available_modes:
                self.set_mode(self.available_modes[0])
            else:
                self.set_mode("EVE")

            self.status.setText(f"Готово. REF: {self.tin_ref_col}, Описание: {self.tin_desc_col}.")
            self._update_connection_buttons(connected=True)
            return True
        except Exception as e:
            self.engine = None
            self.status.setText(f"Ошибка подключения: {e}")
            self._update_connection_buttons(connected=False)
            return False

    def _update_connection_buttons(self, connected: bool):
        # top-right button
        if connected:
            self.btnConn.setText("Подключено")
            self.btnConn.setEnabled(False)
            if hasattr(self, "btnQuickConnect"):
                self.btnQuickConnect.setText("Подключено")
                self.btnQuickConnect.setEnabled(False)
        else:
            self.btnConn.setText("Настроить подключение к БД")
            self.btnConn.setEnabled(True)
            if hasattr(self, "btnQuickConnect"):
                self.btnQuickConnect.setText("Подключиться")
                self.btnQuickConnect.setEnabled(True)

    def _load_modes(self):
        self.available_modes = ["EVE", "S", "F"]
        if self.engine is None:
            return
        try:
            cols = [c['name'] for c in self.inspector.get_columns(MODES_TABLE)]
            if not cols:
                return
            def find(col_names: List[str]) -> Optional[str]:
                s = {n.lower(): n for n in cols}
                for name in col_names:
                    if name.lower() in s: return s[name.lower()]
                for c in cols:
                    if c.lower().replace(" ", "") in {n.lower().replace(" ", "") for n in col_names}:
                        return c
                return None
            mode_col = find([MODES_COL])
            if not mode_col:
                return
            sql = f'''
                SELECT DISTINCT {qident(mode_col)} AS mode
                FROM {qident(MODES_TABLE)}
                WHERE {qident(mode_col)} IS NOT NULL AND trim({qident(mode_col)}::text) <> ''
                ORDER BY {qident(mode_col)}
            '''
            modes: List[str] = []
            with self.engine.connect() as conn:
                for (m,) in conn.execute(text(sql)):
                    s = str(m).strip()
                    if s and s not in modes:
                        modes.append(s)
            if modes:
                self.available_modes = modes
        except Exception:
            pass

    def _detect_tin_columns(self):
        self.tin_ref_col = None
        self.tin_desc_col = None
        self.tin_pack_col = None
        self.tin_price_list_col = None
        self.tin_price_trimm_col = None

        cols = [c['name'] for c in self.inspector.get_columns(TIN_ALL_TABLE)]

        def norm(s): return (s or "").lower().replace(" ", "")

        candidates = ["ref#","ref #","ref","pn","кат. №","кат.#","артикул"]
        for c in cols:
            n = norm(c)
            if n in {x.replace(" ","") for x in candidates} or "ref" in n:
                self.tin_ref_col = c; break
        if not self.tin_ref_col and cols: self.tin_ref_col = cols[0]

        for c in cols:
            if norm(c) in ("наименованиерус","наименование","описание","описаное"):
                self.tin_desc_col = c; break
        if not self.tin_desc_col and len(cols)>1: self.tin_desc_col = cols[1] if len(cols)>1 else cols[0]

        for c in cols:
            if norm(c) in ("вуп-ке","вуп","упаковка","шт/уп"):
                self.tin_pack_col = c; break

        for c in cols:
            nc = norm(c)
            if self.tin_price_list_col is None and ("лист" in nc and "25" in nc):
                self.tin_price_list_col = c
            if self.tin_price_trimm_col is None and (("трм" in nc or "трим" in nc or "трмм" in nc) and "25" in nc):
                self.tin_price_trimm_col = c
        if self.tin_price_trimm_col:
            candidates_trm = [c for c in cols if ("трм" in norm(c) or "трим" in norm(c) or "трмм" in norm(c)) and "25" in norm(c)]
            for c in candidates_trm:
                if "спец" in norm(c):
                    self.tin_price_trimm_col = c
                    break

    def _preload_all_db(self):
        self.table_cache.clear()
        self.options_cache.clear()
        self.tin_index.clear()
        if self.engine is None: return
        with self.engine.connect() as conn:
            for label, table in LABEL_TO_TABLE.items():
                rows = []
                try:
                    cols = has_columns(self.engine, table, ["DIV","Disc Sh","PN","Side","Сторона"])
                    side_expr = None
                    if cols.get("Side"):      side_expr = qident("Side")
                    elif cols.get("Сторона"): side_expr = qident("Сторона")

                    sel = f'SELECT {qident("DIV")}, {qident("Disc Sh")}, {qident("PN")}'
                    sel += f', {side_expr} AS "SIDE"' if side_expr else ', NULL AS "SIDE"'
                    sel += f' FROM {qident(table)}'

                    for row in conn.execute(text(sel)):
                        rows.append({
                            "DIV": (row[0] or "").strip(),
                            "Disc Sh": (row[1] or "").strip(),
                            "PN": norm_ref(row[2]),
                            "SIDE": (row[3] or "").strip() if row[3] is not None else ""
                        })
                except Exception:
                    rows = []
                self.table_cache[table] = rows
                for mode in self.available_modes:
                    for side in (None, "прав", "лев"):
                        self._build_options_cache(table, mode, side)

            ref_col = qident(self.tin_ref_col)
            desc_col = qident(self.tin_desc_col)
            pack_col = qident(self.tin_pack_col) if self.tin_pack_col else "NULL"
            list_col = qident(self.tin_price_list_col) if self.tin_price_list_col else "NULL"
            trm_col  = qident(self.tin_price_trimm_col) if self.tin_price_trimm_col else "NULL"
            sel_tin = f'''
                SELECT {ref_col} AS ref, {desc_col} AS desc_ru, {pack_col} AS pack,
                       {list_col} AS price_list, {trm_col} AS price_trimm
                FROM {qident(TIN_ALL_TABLE)}
            '''
            try:
                for ref, desc_ru, pack, price_list, price_trimm in conn.execute(text(sel_tin)):
                    if ref is None: continue
                    keys = {norm_ref(ref), str(ref).strip(), digits_only(str(ref))}
                    for k in keys:
                        self.tin_index[k] = (
                            str(desc_ru or ""),
                            None if pack is None else str(pack),
                            None if price_list is None else str(price_list),
                            None if price_trimm is None else str(price_trimm),
                        )
            except Exception:
                pass

    def _build_options_cache(self, table: str, mode: str, side: Optional[str]):
        rows = self.table_cache.get(table, [])
        want_mode = (mode or "").strip().lower()
        mapping: Dict[str, str] = {}
        for r in rows:
            div = (r.get("DIV") or "").strip().lower()
            if div != want_mode:
                continue
            if side and r.get("SIDE") and (r["SIDE"].strip().lower() != side.strip().lower()):
                continue
            disc = r.get("Disc Sh",""); pn = r.get("PN","")
            if disc and pn and disc not in mapping:
                mapping[disc] = pn
        self.options_cache[(table, mode, side)] = mapping

    # ============== Templates ==============
    def _preload_templates(self):
        self.templates.clear()
        if self.engine is None: return
        try:
            cols = [c['name'] for c in self.inspector.get_columns(TEMPLATES_TABLE)]
        except Exception:
            self.status.setText(f'Таблица "{TEMPLATES_TABLE}" не найдена — кнопки шаблонов пропущены.')
            return

        def find(col_names: List[str]) -> Optional[str]:
            s = {n.lower(): n for n in cols}
            for name in col_names:
                if name.lower() in s: return s[name.lower()]
            for c in cols:
                if c.lower().replace(" ", "") in {n.lower().replace(" ", "") for n in col_names}:
                    return c
            return None

        type_col = find(["Type", "Тип"])
        pn_col   = find(["PN", "Кат. №", "Артикул", "Ref", "REF", "REF #", "REF#"])
        qts_col  = find(["Qts"])

        if not (type_col and pn_col):
            self.status.setText(f'В таблице "{TEMPLATES_TABLE}" не найдены столбцы Type/PN.')
            return

        sel = f'SELECT {qident(type_col)} AS t, {qident(pn_col)} AS pn'
        if qts_col:
            sel += f', {qident(qts_col)} AS qts'
        sel += f' FROM {qident(TEMPLATES_TABLE)}'

        try:
            with self.engine.connect() as conn:
                for row in conn.execute(text(sel)):
                    t = str(row[0] or "").strip()
                    pn = norm_ref(row[1])
                    if not t or not pn: continue
                    if qts_col:
                        raw_qts = row[2]
                        try: q = int(float(raw_qts)) if raw_qts is not None else 1
                        except Exception: q = 1
                    else:
                        q = 1
                    key = t.lower()
                    self.templates.setdefault(key, []).append((pn, max(1, q)))
        except Exception as e:
            self.status.setText(f'Ошибка чтения "{TEMPLATES_TABLE}": {e}')

    def _on_template_button(self, type_name: str, mode: Optional[str] = None, side: Optional[str] = None):
        if mode:
            self.set_mode(mode, side)
        self._apply_template(type_name)

    def _apply_template(self, type_name: str):
        if not self.templates:
            self._preload_templates()
        items = self.templates.get(type_name.strip().lower(), [])
        if not items:
            QtWidgets.QMessageBox.information(self, "Шаблон", f'В "{TEMPLATES_TABLE}" нет строк с Type = "{type_name}".')
            return
        added = 0
        for pn, qts in items:
            added += self._add_row_by_pn(pn, qts)
        if added:
            self._renumber()
            self.status.setText(f'Добавлено из "{type_name}": {added} позиций (кол-во из Qts).')
        else:
            self.status.setText(f'Шаблон "{type_name}" не дал совпадений по PN.')

    # ============== Interactions ==============
    def set_mode(self, mode: str, side: Optional[str] = None):
        m = (mode or "").strip()
        if m.lower() == "f":
            self.current_mode, self.side_filter = m, side
        else:
            self.current_mode, self.side_filter = m, None
        self._rebuild_left()
        self.status.setText(f"Режим: {self.current_mode}" + (f", {self.side_filter}" if self.side_filter else ""))

       def _fetch_tin_by_pn(self, pn: str):
        if not pn or self.engine is None or not self.tin_ref_col:
            return ("", None, None, None)

        pn_text = str(pn).strip().replace(",", ".")
        pn_digits = digits_only(pn_text)

        pn_num = None
        try:
            f = float(pn_text)
            pn_num = int(f) if f.is_integer() else f
        except Exception:
            pn_num = None

        ref = qident(self.tin_ref_col)
        desc = qident(self.tin_desc_col)
        pack = qident(self.tin_pack_col) if self.tin_pack_col else "NULL"
        list_col = qident(self.tin_price_list_col) if self.tin_price_list_col else "NULL"
        trm_col = qident(self.tin_price_trimm_col) if self.tin_price_trimm_col else "NULL"

        # Build optional numeric clause separately (avoid backslashes inside f-string expressions)
        or_numeric = ""
        if pn_num is not None:
            or_numeric = (
                "       OR ({ref}::text ~ '^\\s*\\d+(\\.\\d+)?\\s*' "
                "AND {ref}::numeric = :pn_num)\n"
            ).format(ref=ref)

        sql = (
            "SELECT {desc} AS desc_ru, {pack} AS pack, {list_col} AS price_list, {trm_col} AS price_trimm\n"
            "FROM {table}\n"
            "WHERE trim({ref}::text) = :pn_text\n"
            "   OR regexp_replace({ref}::text, '\\\\D', '', 'g') = :pn_digits\n"
            "{or_numeric}"
            "LIMIT 1"
        ).format(
            desc=desc,
            pack=pack,
            list_col=list_col,
            trm_col=trm_col,
            table=qident(TIN_ALL_TABLE),
            ref=ref,
            or_numeric=or_numeric,
        )

        try:
            with self.engine.connect() as conn:
                params = {"pn_text": pn_text, "pn_digits": pn_digits}
                if pn_num is not None:
                    params["pn_num"] = pn_num
                row = conn.execute(text(sql), params).fetchone()
                if row:
                    return (
                        str(row[0] or ""),
                        None if row[1] is None else str(row[1]),
                        None if row[2] is None else str(row[2]),
                        None if row[3] is None else str(row[3]),
                    )
        except Exception:
            pass
        return ("", None, None, None)


    def add_to_summary(self, label: str, combo: QtWidgets.QComboBox, spin: QtWidgets.QSpinBox):
        option = combo.currentText()
        qty = max(1, spin.value() or 1)
        pn = ""
        table_name = LABEL_TO_TABLE.get(label)
        if table_name:
            pn_map = self.options_cache.get((table_name, self.current_mode, self.side_filter)) or {}
            pn = pn_map.get(option, "")
        self._add_row_by_pn(pn, qty)
        spin.setValue(0)
        self._renumber()
        self._recalc_totals()

    def _add_row_by_pn(self, pn: str, qty: int = 1) -> int:
        if not pn:
            return 0
        desc, pack, price_list_s, price_trm_s = self._fetch_tin_by_pn(pn)
        if not desc:
            hit = self.tin_index.get(pn) or self.tin_index.get(norm_ref(pn)) or self.tin_index.get(digits_only(pn))
            if hit:
                desc, pack, price_list_s, price_trm_s = hit

        list_price = parse_money(price_list_s)
        trm_base   = parse_money(price_trm_s)
        trm_display = trm_base * self.logisticsSpin.value() if trm_base is not None else None

        self._building_table = True
        try:
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, self.COL_NO, QtWidgets.QTableWidgetItem(str(row+1)))
            self.table.setItem(row, self.COL_PN, QtWidgets.QTableWidgetItem(pn))
            self.table.setItem(row, self.COL_DESC, QtWidgets.QTableWidgetItem(desc))
            self.table.setItem(row, self.COL_QTY, QtWidgets.QTableWidgetItem(str(max(1, qty))))
            self.table.setItem(row, self.COL_CTRL, QtWidgets.QTableWidgetItem("1"))
            self.table.setItem(row, self.COL_PACK, QtWidgets.QTableWidgetItem("" if pack is None else str(pack)))

            it_list = QtWidgets.QTableWidgetItem(fmt_money(list_price))
            it_list.setData(self.USERROLE_LIST_PRICE, list_price if list_price is not None else 0.0)
            self.table.setItem(row, self.COL_PRICE_LIST, it_list)

            it_trm = QtWidgets.QTableWidgetItem(fmt_money(trm_display))
            it_trm.setData(self.USERROLE_BASE_TRIMM, trm_base if trm_base is not None else 0.0)
            self.table.setItem(row, self.COL_PRICE_TRIMM, it_trm)
        finally:
            self._building_table = False
        return 1

    def _renumber(self):
        for r in range(self.table.rowCount()):
            self.table.setItem(r, self.COL_NO, QtWidgets.QTableWidgetItem(str(r+1)))

    def delete_selected(self):
        rows = sorted({ix.row() for ix in self.table.selectedIndexes()}, reverse=True)
        for r in rows:
            self.table.removeRow(r)
        self._renumber()
        self._recalc_totals()

    def _clear_cat_numbers(self):
        col = self.COL_PN
        for r in range(self.table.rowCount()):
            self.table.setItem(r, col, QtWidgets.QTableWidgetItem(""))

    def _move_row_with_cat_down(self):
        col = self.COL_PN
        for r in range(self.table.rowCount()-1):
            it = self.table.item(r, col)
            if it and it.text().strip():
                self.table.insertRow(r+2)
                for c in range(self.table.columnCount()):
                    src = self.table.takeItem(r, c)
                    self.table.setItem(r+2, c, src)
                self.table.removeRow(r)
                break
        self._renumber()
        self._recalc_totals()

    def _on_table_item_changed(self, _: QtWidgets.QTableWidgetItem):
        if self._building_table: return
        self._recalc_totals()

    def _on_logistics_changed(self, _val: float):
        self._building_table = True
        try:
            L = self.logisticsSpin.value()
            for r in range(self.table.rowCount()):
                it_trm = self.table.item(r, self.COL_PRICE_TRIMM)
                if not it_trm: continue
                base = it_trm.data(self.USERROLE_BASE_TRIMM)
                base = float(base) if base is not None else 0.0
                disp = base * L if base else 0.0
                it_trm.setText(fmt_money(disp if base else None))
        finally:
            self._building_table = False
        self._recalc_totals()

    def _recalc_totals(self):
        disc = 1.0 - (self.discountSpin.value() / 100.0)
        L = self.logisticsSpin.value()
        K = self.kursSpin.value()

        list_total = 0.0
        trm_total  = 0.0
        for r in range(self.table.rowCount()):
            try:
                qty = int(float((self.table.item(r, self.COL_QTY).text() or "0").replace(",", ".")))
            except Exception:
                qty = 0
            it_list = self.table.item(r, self.COL_PRICE_LIST)
            list_pu = float(it_list.data(self.USERROLE_LIST_PRICE) or 0.0) if it_list else 0.0
            it_trm = self.table.item(r, self.COL_PRICE_TRIMM)
            trm_base = float(it_trm.data(self.USERROLE_BASE_TRIMM) or 0.0) if it_trm else 0.0

            list_total += qty * list_pu
            trm_total  += qty * trm_base * L

        total_after_disc = list_total * disc
        total_after_disc_conv = total_after_disc * K
        trm_total_conv = trm_total * K

        if trm_total_conv > 0:
            margin = (total_after_disc_conv - trm_total_conv) / trm_total_conv
            margin_text = f"Маржа: {margin*100:.2f}%"
        else:
            margin_text = "Маржа: —"

        self.lbl_total.setText(f"Итого: {total_after_disc_conv:.2f}")
        self.lbl_margin.setText(margin_text)

    # ============== Export KP ==============
    def _export_kp(self):
        if openpyxl is None:
            QtWidgets.QMessageBox.critical(self, "Выгрузить КП",
                "Модуль openpyxl не установлен. Установите его:\n\npip install openpyxl")
            return

        items: List[Tuple[str, str, int, float]] = []
        for r in range(self.table.rowCount()):
            pn = (self.table.item(r, self.COL_PN).text().strip()
                  if self.table.item(r, self.COL_PN) else "")
            desc = (self.table.item(r, self.COL_DESC).text().strip()
                    if self.table.item(r, self.COL_DESC) else "")
            qty_text = (self.table.item(r, self.COL_QTY).text()
                        if self.table.item(r, self.COL_QTY) else "")
            try:
                qty = int(float((qty_text or "0").replace(",", ".")))
            except Exception:
                qty = 0
            it_list = self.table.item(r, self.COL_PRICE_LIST)
            unit_price = float(it_list.data(self.USERROLE_LIST_PRICE) or 0.0) if it_list else 0.0
            if not (pn or desc or qty or unit_price): continue
            items.append((pn, desc, qty, unit_price))

        if not items:
            QtWidgets.QMessageBox.information(self, "Выгрузить КП", "Нет данных в сводке для выгрузки.")
            return

        base_dir = app_dir()
        tpl_path = os.path.join(base_dir, "template.xlsx")
        if not os.path.exists(tpl_path):
            QtWidgets.QMessageBox.critical(self, "Выгрузить КП",
                f"Не найден шаблон: {tpl_path}\nПоложите template.xlsx рядом с приложением.")
            return
        out_path = os.path.join(base_dir, f"КП_{datetime.now():%Y%m%d_%H%M%S}.xlsx")

        try:
            wb = openpyxl.load_workbook(tpl_path, keep_vba=False)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Выгрузить КП", f"Не удалось открыть шаблон:\n{e}")
            return

        ws = wb.active

        def write_line(row_index: int, rec: Tuple[str, str, int, float]):
            pn, desc, qty, unit_price = rec
            ws.cell(row=row_index, column=2, value=pn)
            ws.cell(row=row_index, column=3, value=desc)
            ws.cell(row=row_index, column=4, value=qty)
            ws.cell(row=row_index, column=5, value=unit_price)
            ws.cell(row=row_index, column=6, value=f"=E{row_index}*D{row_index}")

        write_line(5, items[0])
        start_row = 22
        for i, rec in enumerate(items[1:], start=0):
            write_line(start_row + i, rec)

        try:
            wb.save(out_path)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Выгрузить КП", f"Не удалось сохранить файл:\n{e}")
            return

        try:
            if sys.platform.startswith("win"): os.startfile(out_path)
            elif sys.platform == "darwin": subprocess.Popen(["open", out_path])
            else: subprocess.Popen(["xdg-open", out_path])
        except Exception:
            pass

        QtWidgets.QMessageBox.information(self, "Выгрузить КП",
            f"Готово: записано {len(items)} строк(и) в файл:\n{out_path}")

    # ============== Other actions ==============
    def _kp_ivl(self):
        QtWidgets.QMessageBox.information(self, "КП ИВЛ", "Формирование КП ИВЛ.")

    def _kp_raskh(self):
        QtWidgets.QMessageBox.information(self, "КП расх", "Формирование КП: расходники.")

    def _po_stephan(self):
        QtWidgets.QMessageBox.information(self, "PO Stephan", "Оформление PO Stephan.")

    # ============== Connection UI ==============
    def _show_conn_dialog(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Подключение к базе данных")
        layout = QFormLayout(dlg)

        host = QtWidgets.QLineEdit(self.hostEdit.text())
        port = QtWidgets.QLineEdit(self.portEdit.text())
        db   = QtWidgets.QLineEdit(self.dbEdit.text())
        user = QtWidgets.QLineEdit(self.userEdit.text())
        pw   = QtWidgets.QLineEdit(self.pwEdit.text()); pw.setEchoMode(QtWidgets.QLineEdit.Password)

        layout.addRow("Хост", host)
        layout.addRow("Порт", port)
        layout.addRow("База данных", db)
        layout.addRow("Пользователь", user)
        layout.addRow("Пароль", pw)

        btns = QDialogButtonBox()
        btn_connect = QtWidgets.QPushButton("Подключиться")
        btn_close   = QtWidgets.QPushButton("Отмена")
        btns.addButton(btn_connect, QDialogButtonBox.AcceptRole)
        btns.addButton(btn_close,   QDialogButtonBox.RejectRole)
        layout.addRow(btns)

        def on_connect_clicked():
            self.hostEdit.setText(host.text())
            self.portEdit.setText(port.text())
            self.dbEdit.setText(db.text())
            self.userEdit.setText(user.text())
            self.pwEdit.setText(pw.text())
            if self.connect_db():
                btn_connect.setText("Подключено")
                btn_connect.setEnabled(False)

        btn_connect.clicked.connect(on_connect_clicked)
        btn_close.clicked.connect(dlg.reject)
        dlg.exec()

    def _quick_connect(self):
        # uses saved creds; no dialog
        self.connect_db()

    def _launch_admin_tool(self):
        tool = os.path.join(app_dir(), "pg_admin_gui.py")
        if not os.path.exists(tool):
            QtWidgets.QMessageBox.critical(self, "Редактировать значения/таблицы",
                f"Не найден файл:\n{tool}\nСохраните Tk-программу рядом с приложением.")
            return
        env = os.environ.copy()
        env.update({
            "PGHOST": self.hostEdit.text(),
            "PGPORT": self.portEdit.text(),
            "PGDATABASE": self.dbEdit.text(),
            "PGUSER": self.userEdit.text(),
            "PGPASSWORD": self.pwEdit.text(),
        })
        try:
            subprocess.Popen([sys.executable, tool], env=env)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Редактировать значения/таблицы", f"Не удалось запустить редактор:\n{e}")

def main():
    app = QtWidgets.QApplication(sys.argv)
    w = Panel(); w.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
